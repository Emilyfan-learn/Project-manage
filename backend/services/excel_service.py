"""
Service for Excel import/export operations
Rewritten to use only openpyxl (no pandas dependency)
"""
import sqlite3
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.config import settings
from backend.models.wbs import WBSCreate
from backend.services.wbs_service import WBSService


class ExcelService:
    """Service for Excel import/export"""

    def __init__(self):
        self.db_path = str(settings.database_path)
        self.wbs_service = WBSService()

    def _parse_date(self, date_value: Any) -> Optional[str]:
        """Parse date from various formats to YYYY-MM-DD"""
        if date_value is None or date_value == '':
            return None

        try:
            # If already a datetime object
            if isinstance(date_value, datetime):
                return date_value.strftime('%Y-%m-%d')

            # If string, try to parse
            date_str = str(date_value).strip()
            if not date_str or date_str.lower() == 'none':
                return None

            # Try different formats
            for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d %H:%M:%S']:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue

            return None

        except Exception:
            return None

    def _is_empty(self, value: Any) -> bool:
        """Check if a cell value is empty"""
        if value is None:
            return True
        if isinstance(value, str) and value.strip() == '':
            return True
        return False

    def _clean_string(self, value: Any) -> Optional[str]:
        """Clean and convert value to string"""
        if self._is_empty(value):
            return None
        return str(value).strip()

    def _clean_parent_id(self, value: Any) -> Optional[str]:
        """Clean parent_id value"""
        if self._is_empty(value):
            return None
        parent_str = str(value).strip()
        # Handle numeric values like 1.0 -> 1
        if parent_str.endswith('.0'):
            base = parent_str[:-2]
            if base.isdigit():
                return base
        return parent_str if parent_str else None

    def import_wbs_from_excel(self, file_path: str, project_id: str) -> Dict[str, Any]:
        """
        Import WBS items from Excel file
        """
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            # Column mapping (Chinese to English)
            column_map = {
                '項目': 'wbs_id',
                '父項目': 'parent_id',
                '任務說明': 'task_name',
                '單位': 'owner_unit',
                '類別': 'category',
                '預計開始 (原始)': 'original_planned_start',
                '預計結束 (原始)': 'original_planned_end',
                '預計開始 (調整)': 'revised_planned_start',
                '預計結束 (調整)': 'revised_planned_end',
                '開始日期': 'actual_start_date',
                '結束日期': 'actual_end_date',
                '工作天數': 'work_days',
                '實際完成進度': 'actual_progress',
                '狀態': 'status',
                '備註說明': 'notes',
                '內部安排': 'is_internal',
            }

            # Map column indices
            col_indices = {}
            for idx, header in enumerate(headers):
                if header in column_map:
                    col_indices[column_map[header]] = idx

            # Check required columns
            required = ['wbs_id', 'task_name']
            missing = [col for col in required if col not in col_indices]
            if missing:
                return {
                    'success': False,
                    'error': f'Missing required columns: {", ".join(missing)}',
                    'imported': 0,
                    'failed': 0
                }

            # Import rows
            imported = []
            failed = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Get WBS ID
                    wbs_id_idx = col_indices.get('wbs_id')
                    wbs_id = row[wbs_id_idx] if wbs_id_idx is not None else None

                    if self._is_empty(wbs_id):
                        continue

                    # Parse is_internal
                    is_internal_value = False
                    if 'is_internal' in col_indices:
                        val = row[col_indices['is_internal']]
                        if not self._is_empty(val):
                            val_str = str(val).strip().lower()
                            is_internal_value = val_str in ['yes', 'y', 'true', '1', '是', 'v', '✓', 'x']

                    # Build WBS data
                    wbs_data = {
                        'project_id': project_id,
                        'wbs_id': str(wbs_id).strip(),
                        'parent_id': self._clean_parent_id(row[col_indices['parent_id']]) if 'parent_id' in col_indices else None,
                        'task_name': self._clean_string(row[col_indices['task_name']]) or '',
                        'category': self._clean_string(row[col_indices.get('category')]) or 'Task',
                        'owner_unit': self._clean_string(row[col_indices.get('owner_unit')]),
                        'original_planned_start': self._parse_date(row[col_indices.get('original_planned_start')]) if 'original_planned_start' in col_indices else None,
                        'original_planned_end': self._parse_date(row[col_indices.get('original_planned_end')]) if 'original_planned_end' in col_indices else None,
                        'revised_planned_start': self._parse_date(row[col_indices.get('revised_planned_start')]) if 'revised_planned_start' in col_indices else None,
                        'revised_planned_end': self._parse_date(row[col_indices.get('revised_planned_end')]) if 'revised_planned_end' in col_indices else None,
                        'actual_start_date': self._parse_date(row[col_indices.get('actual_start_date')]) if 'actual_start_date' in col_indices else None,
                        'actual_end_date': self._parse_date(row[col_indices.get('actual_end_date')]) if 'actual_end_date' in col_indices else None,
                        'work_days': int(row[col_indices['work_days']]) if 'work_days' in col_indices and not self._is_empty(row[col_indices.get('work_days')]) else None,
                        'actual_progress': int(row[col_indices['actual_progress']]) if 'actual_progress' in col_indices and not self._is_empty(row[col_indices.get('actual_progress')]) else 0,
                        'status': self._clean_string(row[col_indices.get('status')]) or '未開始',
                        'notes': self._clean_string(row[col_indices.get('notes')]),
                        'is_internal': is_internal_value,
                    }

                    # Create WBS item
                    wbs_create = WBSCreate(**wbs_data)
                    self.wbs_service.create_wbs(wbs_create)

                    imported.append({
                        'row': row_idx,
                        'wbs_id': wbs_data['wbs_id'],
                        'task_name': wbs_data['task_name']
                    })

                except Exception as e:
                    failed.append({
                        'row': row_idx,
                        'wbs_id': str(row[col_indices.get('wbs_id', 0)] if col_indices.get('wbs_id') is not None else 'N/A'),
                        'error': str(e)
                    })

            wb.close()

            return {
                'success': len(imported) > 0,
                'imported': len(imported),
                'failed': len(failed),
                'imported_items': imported,
                'failed_items': failed
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'imported': 0,
                'failed': 0
            }

    def export_wbs_to_excel(self, project_id: str, output_path: str) -> Dict[str, Any]:
        """
        Export WBS items to Excel file
        """
        try:
            # Get WBS data from database
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    wbs_id, parent_id, task_name, owner_unit, category,
                    original_planned_start, original_planned_end,
                    revised_planned_start, revised_planned_end,
                    actual_start_date, actual_end_date,
                    work_days, actual_progress, estimated_progress, progress_variance,
                    status, notes, is_internal, is_overdue
                FROM tracking_items
                WHERE project_id = ? AND item_type = 'WBS'
                ORDER BY wbs_id
            """, (project_id,))

            rows = cursor.fetchall()
            conn.close()

            if not rows:
                return {
                    'success': False,
                    'error': 'No WBS items found for this project',
                    'exported': 0
                }

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'WBS'

            # Headers
            headers = [
                '項目', '父項目', '任務說明', '單位', '類別',
                '預計開始 (原始)', '預計結束 (原始)',
                '預計開始 (調整)', '預計結束 (調整)',
                '開始日期', '結束日期',
                '工作天數', '實際完成進度', '預估完成進度', '進度偏差',
                '狀態', '備註說明', '內部安排', '逾期'
            ]

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Write data
            overdue_rows = []
            for row_idx, row in enumerate(rows, 2):
                data = [
                    row['wbs_id'],
                    row['parent_id'],
                    row['task_name'],
                    row['owner_unit'],
                    row['category'],
                    row['original_planned_start'],
                    row['original_planned_end'],
                    row['revised_planned_start'],
                    row['revised_planned_end'],
                    row['actual_start_date'],
                    row['actual_end_date'],
                    row['work_days'],
                    row['actual_progress'],
                    row['estimated_progress'],
                    row['progress_variance'],
                    row['status'],
                    row['notes'],
                    'V' if row['is_internal'] else '',
                    '是' if row['is_overdue'] else '否'
                ]
                for col, value in enumerate(data, 1):
                    ws.cell(row=row_idx, column=col, value=value)

                if row['is_overdue']:
                    overdue_rows.append(row_idx)

            # Apply styles
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            overdue_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')

            for row in ws.iter_rows(min_row=1, max_row=len(rows) + 1, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = border

            # Highlight overdue rows
            for row_idx in overdue_rows:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = overdue_fill

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                max_length = len(str(headers[col - 1]))
                for row in range(2, len(rows) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

            wb.save(output_path)
            wb.close()

            return {
                'success': True,
                'exported': len(rows),
                'file_path': output_path
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'exported': 0
            }

    def create_wbs_template(self, output_path: str) -> Dict[str, Any]:
        """
        Create WBS import template Excel file
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'WBS範本'

            # Headers
            headers = [
                '項目', '父項目', '任務說明', '單位', '類別',
                '預計開始 (原始)', '預計結束 (原始)',
                '預計開始 (調整)', '預計結束 (調整)',
                '開始日期', '結束日期',
                '工作天數', '實際完成進度', '狀態', '內部安排', '備註說明'
            ]

            # Sample data
            sample_data = [
                ['1', '', '專案啟動', '專案經理', 'Milestone', '2024/01/01', '2024/01/01', '', '', '', '', '', 100, '已完成', '', '頂層項目範例'],
                ['1.1', '1', '需求分析', '開發部', 'Task', '2024/01/02', '2024/01/15', '', '', '2024/01/02', '2024/01/14', 10, 100, '已完成', '', '子項目範例'],
                ['1.2', '1', '系統設計', 'AAA/BBB', 'Task', '2024/01/16', '2024/01/31', '', '', '2024/01/16', '', 12, 60, '進行中', 'V', '子項目範例（內部安排）'],
                ['2', '', '開發階段', '開發部', 'Milestone', '2024/02/01', '2024/03/31', '', '', '', '', '', 0, '未開始', '', '頂層項目範例'],
            ]

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Write sample data
            for row_idx, row_data in enumerate(sample_data, 2):
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=value)

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                max_length = len(str(headers[col - 1]))
                for row in range(2, len(sample_data) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 30)

            wb.save(output_path)
            wb.close()

            return {
                'success': True,
                'file_path': output_path,
                'message': 'Template created successfully'
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    def export_pending_to_excel(self, project_id: str, output_path: str) -> Dict[str, Any]:
        """
        Export Pending items to Excel file
        """
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    pending_id, task_date, source_type, contact_info, description,
                    expected_reply_date, is_replied, actual_reply_date,
                    handling_notes, related_wbs, related_action_item,
                    status, priority
                FROM pending_items
                WHERE project_id = ?
                ORDER BY task_date DESC
            """, (project_id,))

            rows = cursor.fetchall()
            conn.close()

            if not rows:
                return {
                    'success': False,
                    'error': 'No pending items found for this project',
                    'exported': 0
                }

            wb = Workbook()
            ws = wb.active
            ws.title = '待辦事項'

            headers = [
                '編號', '任務日期', '來源類型', '聯絡資訊', '說明',
                '預計回覆日期', '已回覆', '實際回覆日期',
                '處理備註', '相關 WBS', '相關行動項目', '狀態', '優先級'
            ]

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Write data
            for row_idx, row in enumerate(rows, 2):
                data = [
                    row['pending_id'],
                    row['task_date'],
                    row['source_type'],
                    row['contact_info'],
                    row['description'],
                    row['expected_reply_date'],
                    '是' if row['is_replied'] else '否',
                    row['actual_reply_date'],
                    row['handling_notes'],
                    row['related_wbs'],
                    row['related_action_item'],
                    row['status'],
                    row['priority']
                ]
                for col, value in enumerate(data, 1):
                    ws.cell(row=row_idx, column=col, value=value)

            # Apply styles and adjust widths
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1, max_row=len(rows) + 1):
                for cell in row:
                    cell.border = border

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

            wb.save(output_path)
            wb.close()

            return {
                'success': True,
                'exported': len(rows),
                'file_path': output_path
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'exported': 0
            }

    def export_issues_to_excel(self, project_id: str, output_path: str) -> Dict[str, Any]:
        """
        Export Issues to Excel file
        """
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    issue_number, issue_title, issue_description,
                    issue_type, issue_category, severity, priority,
                    reported_by, reported_date, assigned_to, owner_type,
                    affected_wbs, impact_description, estimated_impact_days,
                    status, resolution, root_cause,
                    target_resolution_date, actual_resolution_date, closed_date,
                    is_escalated, escalation_level, escalation_date, escalation_reason
                FROM issue_tracking
                WHERE project_id = ?
                ORDER BY issue_number
            """, (project_id,))

            rows = cursor.fetchall()
            conn.close()

            if not rows:
                return {
                    'success': False,
                    'error': 'No issues found for this project',
                    'exported': 0
                }

            wb = Workbook()
            ws = wb.active
            ws.title = '問題追蹤'

            headers = [
                '問題編號', '問題標題', '問題說明',
                '問題類型', '問題分類', '嚴重性', '優先級',
                '回報人', '回報日期', '指派給', '負責類型',
                '受影響 WBS', '影響說明', '預估影響天數',
                '狀態', '解決方案', '根本原因',
                '目標解決日期', '實際解決日期', '關閉日期',
                '已升級', '升級層級', '升級日期', '升級原因'
            ]

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Write data
            escalated_rows = []
            for row_idx, row in enumerate(rows, 2):
                data = [
                    row['issue_number'], row['issue_title'], row['issue_description'],
                    row['issue_type'], row['issue_category'], row['severity'], row['priority'],
                    row['reported_by'], row['reported_date'], row['assigned_to'], row['owner_type'],
                    row['affected_wbs'], row['impact_description'], row['estimated_impact_days'],
                    row['status'], row['resolution'], row['root_cause'],
                    row['target_resolution_date'], row['actual_resolution_date'], row['closed_date'],
                    '是' if row['is_escalated'] else '否',
                    row['escalation_level'], row['escalation_date'], row['escalation_reason']
                ]
                for col, value in enumerate(data, 1):
                    ws.cell(row=row_idx, column=col, value=value)

                if row['is_escalated']:
                    escalated_rows.append(row_idx)

            # Apply styles
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            escalated_fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')

            for row in ws.iter_rows(min_row=1, max_row=len(rows) + 1):
                for cell in row:
                    cell.border = border

            for row_idx in escalated_rows:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = escalated_fill

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

            wb.save(output_path)
            wb.close()

            return {
                'success': True,
                'exported': len(rows),
                'file_path': output_path
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'exported': 0
            }
